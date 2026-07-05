import openpyxl
from datetime import datetime
from database.db_session import SessionLocal
from database.models.animal import Animal
from controllers.animal_controller import AnimalController

def importer_animaux_depuis_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    db = SessionLocal()
    lignes_importees = 0
    erreurs = []

    # Détection des colonnes
    headers = [cell.value for cell in ws[1]]
    col_idx = {}
    mapping = {
        'code_elevage': ['code_elevage', 'code élevage'],
        'numero_boucle': ['numero_boucle', 'numéro boucle', 'boucle'],
        'date_naissance': ['date_naissance', 'date naissance'],
        'sexe': ['sexe'],
        'race': ['race'],
        'origine': ['origine'],
        'date_entree': ['date_entree', 'date entrée'],
        'categorie': ['categorie']
    }
    for i, h in enumerate(headers):
        if h:
            h_lower = str(h).lower().replace(" ", "")
            for key, variants in mapping.items():
                if any(var.lower().replace(" ", "") == h_lower for var in variants):
                    col_idx[key] = i
                    break

    required = ['code_elevage', 'numero_boucle', 'date_naissance', 'sexe', 'race', 'origine']
    missing = [r for r in required if r not in col_idx]
    if missing:
        return 0, f"Colonnes manquantes pour les animaux : {missing}"

    for row in ws.iter_rows(min_row=2, values_only=True):
        code_elevage = row[col_idx['code_elevage']]
        if not code_elevage:
            continue
        numero_boucle = str(row[col_idx['numero_boucle']]).strip()
        if not numero_boucle:
            erreurs.append(f"Numéro de boucle manquant pour {code_elevage}")
            continue
        sexe = row[col_idx['sexe']]
        race = row[col_idx['race']]
        origine = row[col_idx['origine']]
        date_entree_val = row[col_idx['date_entree']] if col_idx.get('date_entree') is not None else None
        categorie_importee = row[col_idx['categorie']] if col_idx.get('categorie') is not None else None

        # Date naissance
        date_val = row[col_idx['date_naissance']]
        try:
            if isinstance(date_val, datetime):
                date_naissance = date_val.strftime("%Y-%m-%d")
            else:
                date_naissance = datetime.strptime(str(date_val), "%d/%m/%Y").strftime("%Y-%m-%d")
        except:
            erreurs.append(f"Date naissance invalide pour {numero_boucle} : {date_val}")
            continue

        # Date entrée si externe
        if origine == "Externe" and date_entree_val:
            try:
                if isinstance(date_entree_val, datetime):
                    date_entree_str = date_entree_val.strftime("%Y-%m-%d")
                else:
                    date_entree_str = datetime.strptime(str(date_entree_val), "%d/%m/%Y").strftime("%Y-%m-%d")
            except:
                date_entree_str = ""
        else:
            date_entree_str = ""

        # Unicité boucle
        existant = db.query(Animal).filter(
            Animal.code_elevage == code_elevage,
            Animal.numero_boucle == numero_boucle
        ).first()
        if existant:
            erreurs.append(f"Boucle {numero_boucle} déjà existante pour l'éleveur {code_elevage}")
            continue

        # Catégorie
        if categorie_importee:
            categorie = categorie_importee
        else:
            categorie = AnimalController.calculer_categorie(sexe, date_naissance)

        animal = Animal(
            code_elevage=code_elevage,
            numero_boucle=numero_boucle,
            date_naissance=date_naissance,
            sexe=sexe,
            race=race,
            origine=origine,
            date_entree=date_entree_str,
            categorie=categorie,
            actif=True
        )
        db.add(animal)
        lignes_importees += 1

    db.commit()
    db.close()
    return lignes_importees, erreurs