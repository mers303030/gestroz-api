import openpyxl
from datetime import datetime
from database.db_session import SessionLocal
from database.models.eleveur import Eleveur

def importer_eleveurs_depuis_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    db = SessionLocal()
    lignes_importees = 0
    erreurs = []

    # Lecture de la première ligne pour trouver les colonnes
    headers = [cell.value for cell in ws[1]]
    col_idx = {}
    mapping = {
        'code_elevage': ['code_elevage', 'code élevage'],
        'nom': ['nom', 'Nom'],
        'prenom': ['prenom', 'Prenom'],
        'date_naissance': ['date_de_naissance', 'Date_de_naissance'],
        'niveau_scolaire': ['niveau_scolaire', 'Niveau_scolaire'],
        'cnie': ['cine', 'CINE'],
        'telephone': ['numero_de_telephone', 'Numero_de_telephone'],
        'commune': ['communerurale', 'CommuneRurale']
    }
    for i, h in enumerate(headers):
        if h:
            h_lower = str(h).lower().replace(" ", "")
            for key, variants in mapping.items():
                if any(var.lower().replace(" ", "") == h_lower for var in variants):
                    col_idx[key] = i
                    break

    required = ['code_elevage', 'nom', 'prenom', 'date_naissance', 'cnie', 'commune']
    missing = [r for r in required if r not in col_idx]
    if missing:
        return 0, f"Colonnes manquantes : {missing}"

    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[col_idx['code_elevage']]
        if not code:
            continue

        nom = row[col_idx['nom']]
        prenom = row[col_idx['prenom']]
        niveau = row[col_idx['niveau_scolaire']] if col_idx.get('niveau_scolaire') is not None else ''
        cnie = str(row[col_idx['cnie']]).strip() if row[col_idx['cnie']] else ''
        tel_raw = row[col_idx['telephone']] if col_idx.get('telephone') is not None else ''
        commune_raw = row[col_idx['commune']]

        # Nettoyage téléphone
        tel = ''.join(filter(str.isdigit, str(tel_raw))) if tel_raw else ''
        if tel and len(tel) != 10:
            erreurs.append(f"Téléphone invalide pour {code} : {tel_raw}")
            continue

        # Normalisation commune
        commune = "Boukchmir" if commune_raw == "Bouikchmir" else commune_raw

        # Conversion date
        date_val = row[col_idx['date_naissance']]
        try:
            if isinstance(date_val, datetime):
                date_naissance = date_val.strftime("%Y-%m-%d")
            else:
                date_naissance = datetime.strptime(str(date_val), "%d/%m/%Y").strftime("%Y-%m-%d")
        except:
            erreurs.append(f"Date invalide pour {code} : {date_val}")
            continue

        # Vérifier unicité CNIE et téléphone
        if db.query(Eleveur).filter(Eleveur.cnie == cnie).first():
            erreurs.append(f"CNIE {cnie} déjà existant pour {code}")
            continue
        if tel and db.query(Eleveur).filter(Eleveur.telephone == tel).first():
            erreurs.append(f"Téléphone {tel} déjà existant pour {code}")
            continue

        # Création ou mise à jour
        eleveur = db.query(Eleveur).filter(Eleveur.code_elevage == code).first()
        if eleveur:
            eleveur.nom = nom
            eleveur.prenom = prenom
            eleveur.date_naissance = date_naissance
            eleveur.niveau_scolaire = niveau
            eleveur.cnie = cnie
            eleveur.telephone = tel
            eleveur.commune = commune
        else:
            eleveur = Eleveur(
                code_elevage=code,
                nom=nom,
                prenom=prenom,
                date_naissance=date_naissance,
                niveau_scolaire=niveau,
                cnie=cnie,
                telephone=tel,
                commune=commune
            )
            db.add(eleveur)
        lignes_importees += 1

    db.commit()
    db.close()
    return lignes_importees, erreurs