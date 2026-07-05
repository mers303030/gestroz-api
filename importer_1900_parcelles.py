from openpyxl import load_workbook
import sqlite3

# 1. Ouvrir le fichier Excel
wb = load_workbook("test_1900_parcelles.xlsx")
ws = wb.active

# 2. Connexion à la base
conn = sqlite3.connect("data/zaer.db")
cursor = conn.cursor()

# 3. Récupérer les codes élevage existants
existing_codes = {row[0] for row in cursor.execute("SELECT code_elevage FROM eleveurs").fetchall()}
print(f"📋 {len(existing_codes)} codes élevage trouvés.")

# 4. (Optionnel) Supprimer les anciennes parcelles ?
rep = input("Voulez-vous supprimer toutes les parcelles existantes avant import ? (o/N) : ").strip().lower()
if rep == "o":
    cursor.execute("DELETE FROM parcelles")
    conn.commit()
    print("🗑️ Anciennes parcelles supprimées.")

# 5. Parcourir les lignes
inserted = 0
errors = 0
skipped_codes = set()

for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    code, nom, surface, statut, occupation, obs = row

    # Vérifier que le code existe
    if code not in existing_codes:
        skipped_codes.add(code)
        errors += 1
        continue

    # Vérifier les champs obligatoires
    if not nom or not surface or surface <= 0:
        errors += 1
        continue

    try:
        cursor.execute("""
            INSERT INTO parcelles (code_elevage, numero_parcelle, surface_ha, statut_foncier, occupation_actuelle, observations)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (code, nom, surface, statut or None, occupation or None, obs or None))
        inserted += 1
    except Exception as e:
        errors += 1
        print(f"❌ Erreur ligne {row}: {e}")

# 6. Commit et fermeture
conn.commit()
conn.close()

print(f"\n✅ {inserted} parcelles importées avec succès.")
print(f"⚠️ {errors} lignes ignorées.")
if skipped_codes:
    print(f"   Codes ignorés (inexistants) : {', '.join(sorted(skipped_codes))[:100]}...")